import io
import calendar
import datetime
import streamlit as st
from openpyxl import load_workbook
import pypdfium2 as pdfium

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, HRFlowable

# Optional imports for right-to-left Arabic shaping
try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    ARABIC_SUPPORT = True
except ImportError:
    ARABIC_SUPPORT = False

# ==============================================================================
# DEFAULT CONFIGURATION
# ==============================================================================
MAIN_HEADER_TITLE = "Quran Memorization Sheet"
NUMBER_COL_WIDTH = 25  # Width of the '#' sequence number column (in points)
# ==============================================================================

st.set_page_config(page_title="Attendance PDF Generator", layout="centered")

st.title("📋 Attendance Sheet Generator")
st.write("Upload your student Excel file, select a sheet, and generate customized PDF attendance sheets.")

def reshape_text(text):
    """Reshapes text for proper display in ReportLab."""
    if ARABIC_SUPPORT and any('\u0600' <= c <= '\u06FF' for c in text):
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    return text

# Sidebar Configuration Controls
st.sidebar.header("Document & Layout Settings")
doc_title = st.sidebar.text_input("Main Header Title", value=MAIN_HEADER_TITLE)
max_students = st.sidebar.number_input("Max Students per Page", min_value=5, max_value=25, value=12, step=1)
student_col_width = st.sidebar.number_input("Student Name Column Width (pt)", min_value=60, max_value=250, value=120, step=5)

st.sidebar.markdown("---")
st.sidebar.header("Date Settings")

# Date Selection Controls
current_year = datetime.datetime.now().year
selected_year = st.sidebar.number_input("Year", min_value=2020, max_value=2035, value=current_year, step=1)

month_names_list = list(calendar.month_name)[1:] # ['January', 'February', ...]
start_month_name = st.sidebar.selectbox("Start Month", options=month_names_list, index=7) # Default: August
start_month_idx = month_names_list.index(start_month_name) + 1

num_months = st.sidebar.number_input("Number of Months to Include", min_value=1, max_value=12, value=4, step=1)

days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
selected_day_name = st.sidebar.selectbox("Day of Week for Attendance", options=days_of_week, index=6) # Default: Sunday
selected_day_idx = days_of_week.index(selected_day_name) # 0 = Mon, ..., 6 = Sun

uploaded_file = st.file_uploader("Choose an Excel file (.xlsx)", type=["xlsx"])

def get_sheet_names(file_bytes):
    """Returns a list of all sheet names in the workbook."""
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    return wb.sheetnames

def parse_students(file_bytes, target_sheets):
    """Extracts student names from specified sheet(s)."""
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    all_students = {}
    
    sheets_to_process = wb.sheetnames if "All Sheets" in target_sheets else target_sheets
    
    for sheet_name in sheets_to_process:
        sheet = wb[sheet_name]
        students = []
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            if row[0] is not None and str(row[0]).strip() != "":
                students.append(str(row[0]).strip())
        if students:
            all_students[sheet_name] = students
            
    return all_students

def get_target_months(start_year, start_month, month_count):
    """Calculates a list of (year, month) tuples starting from start_year/start_month."""
    target_months = []
    y = start_year
    m = start_month
    for _ in range(month_count):
        target_months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return target_months

def generate_pdf(students_by_sheet, title_text, max_students_per_page, student_name_width, year, start_month, month_count, target_weekday, preview_only=False):
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=A4,
        leftMargin=28,
        rightMargin=28,
        topMargin=20,
        bottomMargin=20
    )

    story = []
    
    # Calculate months sequence
    months = get_target_months(year, start_month, 1 if preview_only else month_count)

    styles = getSampleStyleSheet()
    
    # Header styles
    header_title_style = ParagraphStyle(
        'MainHeader', parent=styles['Heading1'], fontName='Helvetica-Bold',
        fontSize=20, leading=24, textColor=colors.HexColor('#000000'),
        alignment=1, spaceAfter=6
    )
    
    month_banner_style = ParagraphStyle(
        'MonthBanner', parent=styles['Heading2'], fontName='Helvetica-Bold',
        fontSize=16, leading=20, textColor=colors.HexColor('#000000'),
        alignment=1
    )

    attendance_banner_style = ParagraphStyle(
        'AttBanner', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=12, leading=14, textColor=colors.HexColor('#000000'),
        alignment=1
    )

    student_style = ParagraphStyle(
        'StudentName', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=10, leading=12, textColor=colors.HexColor('#000000'),
        wordWrap='CJK'
    )
    
    header_date_style = ParagraphStyle(
        'HeaderDate', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=11, leading=13, textColor=colors.HexColor('#000000'), alignment=1
    )

    num_col_style = ParagraphStyle(
        'NumCol', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=10, leading=12, textColor=colors.HexColor('#000000'), alignment=1
    )

    printable_width = 538
    printable_height = 620
    header_row_height = 25
    student_row_height = (printable_height - (header_row_height * 2)) / max_students_per_page

    formatted_title = reshape_text(title_text)

    for sheet_name, students in students_by_sheet.items():
        student_chunks = [
            students[i:i + max_students_per_page] 
            for i in range(0, len(students), max_students_per_page)
        ]

        for y, m in months:
            month_name = calendar.month_name[m]
            
            # Find all dates in month 'm' that match target_weekday
            matching_dates = []
            cal = calendar.monthcalendar(y, m)
            for week in cal:
                day = week[target_weekday]
                if day != 0:
                    matching_dates.append(datetime.date(y, m, day))

            num_dates = len(matching_dates)
            if num_dates == 0:
                continue

            for chunk_idx, chunk in enumerate(student_chunks):
                # 1. Main Title
                story.append(Paragraph(formatted_title, header_title_style))
                story.append(HRFlowable(width="100%", thickness=1, color=colors.black, spaceAfter=8, spaceBefore=4))

                # 2. Light Green Month Banner Row
                month_str = f"{month_name} {y}"
                banner_table = Table([[Paragraph(month_str, month_banner_style)]], colWidths=[printable_width])
                banner_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D1E7DD')),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#000000')),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(banner_table)
                story.append(HRFlowable(width="100%", thickness=0, spaceAfter=8))

                # 3. Main Table Header with Custom Dates
                row0 = [
                    Paragraph("<b>#</b>", header_date_style),
                    Paragraph("<b>Student<br/>Name</b>", header_date_style),
                    Paragraph("Attendance + Notes", attendance_banner_style)
                ] + [""] * (num_dates - 1)
                
                row1 = ["", ""] + [Paragraph(f"<b>{d.day} {month_name[:3]}</b>", header_date_style) for d in matching_dates]
                
                table_data = [row0, row1]

                # Populate Rows
                start_student_num = (chunk_idx * max_students_per_page) + 1
                for idx, student_name in enumerate(chunk):
                    student_num_str = str(start_student_num + idx)
                    table_data.append([
                        Paragraph(student_num_str, num_col_style),
                        Paragraph(student_name, student_style)
                    ] + [""] * num_dates)

                # Pad remaining empty rows if chunk < max_students_per_page
                for _ in range(max_students_per_page - len(chunk)):
                    table_data.append(["", ""] + [""] * num_dates)

                # Column Width Calculations
                remaining_width = printable_width - NUMBER_COL_WIDTH - student_name_width
                day_col_width = remaining_width / num_dates if num_dates > 0 else remaining_width
                col_widths = [NUMBER_COL_WIDTH, student_name_width] + [day_col_width] * num_dates
                
                row_heights = [header_row_height, header_row_height] + [student_row_height] * max_students_per_page

                t = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
                t.setStyle(TableStyle([
                    ('SPAN', (0, 0), (0, 1)),
                    ('SPAN', (1, 0), (1, 1)),
                    ('SPAN', (2, 0), (-1, 0)),
                    ('BACKGROUND', (2, 0), (-1, 0), colors.HexColor('#CCCCCC')),
                    ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#000000')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ]))

                story.append(t)
                
                if preview_only:
                    break
                story.append(PageBreak())

            if preview_only:
                break

    if story and isinstance(story[-1], PageBreak):
        story.pop()

    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer

def render_preview_image(pdf_bytes):
    """Renders PDF page 1 to an image for UI preview."""
    pdf = pdfium.PdfDocument(pdf_bytes)
    page = pdf[0]
    image = page.render(scale=2).to_pil()
    img_byte_arr = io.BytesIO()
    image.save(img_byte_arr, format='PNG')
    return img_byte_arr.getvalue()

if uploaded_file is not None:
    file_bytes = uploaded_file.getvalue()
    sheet_names = get_sheet_names(file_bytes)
    
    # Sheet Selection Box
    selected_sheets = st.sidebar.multiselect(
        "Select Sheet(s) to Process",
        options=["All Sheets"] + sheet_names,
        default=[sheet_names[0]]
    )

    if selected_sheets:
        students_by_sheet = parse_students(file_bytes, selected_sheets)
        
        if students_by_sheet:
            st.subheader("🖼️ Page 1 Layout Preview")
            
            preview_pdf = generate_pdf(
                students_by_sheet, 
                doc_title, 
                max_students, 
                student_col_width,
                selected_year,
                start_month_idx,
                num_months,
                selected_day_idx,
                preview_only=True
            )
            preview_img = render_preview_image(preview_pdf)
            st.image(preview_img, caption="Live Layout Preview", use_container_width=True)
            
            st.markdown("---")
            if st.button("🚀 Generate Full PDF"):
                with st.spinner("Compiling PDF..."):
                    full_pdf = generate_pdf(
                        students_by_sheet, 
                        doc_title, 
                        max_students, 
                        student_col_width, 
                        selected_year,
                        start_month_idx,
                        num_months,
                        selected_day_idx,
                        preview_only=False
                    )
                    st.success("PDF generated successfully!")
                    st.download_button(
                        label="📥 Download Complete PDF",
                        data=full_pdf,
                        file_name="Attendance_Sheets.pdf",
                        mime="application/pdf"
                    )
        else:
            st.error("No student names were found in Column A of the selected sheet(s).")